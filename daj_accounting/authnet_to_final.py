from openpyxl import load_workbook

#program will not catch:
# a double last name, or empty name cells

#input('Type File Name To Extract From: ')
#june_auth_gtf.xlsx or july_auth_gtf.xlsx
authnet_file = 'authnet_july.xlsx' 
#do you want to type the member file name?
#'final_gtf.xlsx'
#input('Where Is The Member File? ')
member_file = 'final_members_test.xlsx'

wb = load_workbook(authnet_file)
wb2 = load_workbook(member_file)

ws = wb[wb.sheetnames[0]]
ws2 = wb2[wb2.sheetnames[0]]

def get_total():
    total = 0
    for row in range(1, ws.max_row+1):
        for col in 'B':
            cell = ws["{}{}".format(col, row)].value
        for col in 'C':
            value = ws["{}{}".format(col, row)].value
            if cell == 'Credited':
                total -= float(value)
            if cell == 'Settled Successfully':
                total += float(value)
    print('TOTAL: {0}'.format(total))

def get_column_t():
    hunda = 0
    montha = 0
    inita = 0
    
    for row in range(1, ws.max_row+1):
        for col in 'B':
            cell = ws["{}{}".format(col, row)].value
        for col in 'C':
            value = ws["{}{}".format(col, row)].value
            if cell == 'Credited':
                if value >= 1000:
                    inita -= value
                elif (value == 100 or value == 200) and value < 1000:
                    montha -= value
                else:
                    hunda -= value
            if cell == 'Settled Successfully':
                if value >= 1000:
                    inita += value
                elif (value == 100 or value == 200) and value < 1000:
                    montha += value
                else:
                    hunda += value
    print('initial total: ', inita)
    print('monthly total: ', montha)
    print('hundreds total: ', hunda)

get_total()    
get_column_t()
def peeps_all():
    print('Please Check The Following Cells')
    print('--------------')
    peeps_set = set([])
    
    for row in range(1, ws.max_row+1):
        for col in 'Y':
            f_name = ws["{}{}".format(col, row)].value
            cell_fname = ws["{}{}".format(col, row)]
        for col in 'Z':
            l_name = ws["{}{}".format(col, row)].value
            cell_lname = ws["{}{}".format(col, row)]
        for col in 'B':
            cell = ws["{}{}".format(col, row)].value
            if cell == 'Credited' or cell == 'Settled Successfully':
                if f_name == None and l_name == None:
                    f_name = 'none' + str(cell_fname.value)
                    l_name = 'none' + str(cell_lname.value)
                    new_person = str(f_name + ' ' + l_name)
                    peeps_set.add(new_person)
                    print('empty cells', cell_lname)
                elif f_name == None:
                    f_name = 'none'
                    new_person = str(f_name + ' ' + l_name)
                    peeps_set.add(new_person)
                elif l_name == None:
                    l_name = 'none'
                    new_person = str(f_name + ' ' + l_name)
                    peeps_set.add(new_person)
                    print('empty cells', cell_lname)
                elif len(l_name.split(' ')) == 2:
                    peeps_set.add(f_name + ' ' + l_name)                    
                    last_split = l_name.split(' ')
                    if last_split[0].lower() != 'mc' and \
                    last_split[0].lower() != 'van' and \
                    last_split[0].lower() != 'st' and \
                    last_split[1].lower() != 'ii' and \
                    last_split[1].lower() != 'iii' and \
                    last_split[1].lower() != 'jr.':
                        print('||' + l_name+ '||', cell_lname)
                elif len(l_name.split(' ')) >= 3:
                    print(l_name, cell_lname)
                else:    
                    new_person = str(f_name + ' ' + l_name)
                    peeps_set.add(new_person)
    return peeps_set
    
class Person:
    def __init__(self, first_name, last_name, initial = 0, monthly = 0, \
                                               hundreds = 0, cell = None):
        if first_name == None:
            first_name = 'none'
        if last_name == None:
            last_name = 'none'

        self.first_name = first_name.lower()
        self.last_name = last_name.lower()
        self.initial = initial
        self.monthly = monthly
        self.hundreds = hundreds
        self.cell = cell
        
peeps_names = {person for person in peeps_all()}

def unique_to_person():
    unique = set([])    
    for person in peeps_names:
        new_p = person.split(' ')
        if len(new_p) == 2:
            unique.add(Person(new_p[0], new_p[1]))
        elif len(new_p) == 3:
            if new_p[2].lower() == 'jr.':#jr with last name
                unique.add(Person(new_p[0], new_p[1] + ' ' + new_p[2]))
            elif new_p[2].lower() == 'ii' or new_p[2].lower() == 'iii':#ii or iii
                unique.add(Person(new_p[0], new_p[1] + ' ' + new_p[2]))
            elif new_p[1].lower() == 'van' or new_p[1].lower() == 'mc' \
                            or new_p[1].lower() == 'st':#misc last name combos
                unique.add(Person(new_p[0], new_p[1] + ' ' + new_p[2]))                                      
            elif new_p[2] == 'none':#check for 'none' as false last name
                unique.add(Person(new_p[0], new_p[1]))
            else:#assume a middle name
                unique.add(Person(new_p[0] + ' ' + new_p[1], new_p[2]))
        elif len(new_p) == 4:
            if new_p[3] == 'none':#middle name with false last name as 'none'
                unique.add(Person(new_p[0] + ' ' + new_p[1], new_p[2]))
            elif new_p[1] == '&':#couples
                unique.add(Person(new_p[0] + ' & ' + new_p[2], new_p[3]))
            else:
                unique.add(Person(new_p[0] + ' ' + new_p[1] + ' ' + new_p[2], \
                                                                   new_p[3]))                     
        else:
            unique.add(Person(new_p[0], new_p[len(new_p)]))
    return unique

def get_amounts():
    peeps = unique_to_person()
    
    for row in range(1, ws.max_row+1):
        for col in 'Y':
            if ws["{}{}".format(col, row)].value == None:
                ws["{}{}".format(col, row)].value = 'none'
            else:
                f_name = ws["{}{}".format(col, row)].value.lower()
        for col in 'Z':
            if ws["{}{}".format(col, row)].value == None:
                ws["{}{}".format(col, row)].value = 'none'
            else:
                l_name = ws["{}{}".format(col, row)].value.lower()
        for col in 'B':
            cell = ws["{}{}".format(col, row)].value
        for col in 'C':
            value = ws["{}{}".format(col, row)].value            
            if cell == 'Settled Successfully':
                for person in peeps:                        
                    if person.first_name == f_name and person.last_name == l_name:
                        if value >= 1000:
                            person.initial += value
                        elif value == 100 or value == 200:
                            person.hundreds += value
                        elif (value != 100 or value != 200) and value < 1000:
                            person.monthly += value
            elif cell == 'Credited':                
                for person in peeps:
                    if person.first_name == f_name and person.last_name == l_name:                        
                        if value >= 1000:
                            person.initial -= value
                        elif value == 100 or value == 200:
                            person.hundreds -= value
                        elif (value != 100 or value != 200) and value < 1000:
                            person.monthly -= value
    return peeps                        

peeps = get_amounts()

def map_to_final():
    members_set = set([])
    
    for row in range(2, ws2.max_row+1):
        for col in 'G':
            initial = "{}{}".format(col, row)
        for col in 'H':
            monthly = "{}{}".format(col, row)
        for col in 'I':
            hundreds = "{}{}".format(col, row)
        for col in 'A':
            if isinstance(ws2["{}{}".format(col, row)].value, str):
                f_name = ws2["{}{}".format(col, row)].value.lower()
            else:
                l_name = ws2["{}{}".format(col, row)].value = 'none'
        for col in 'B':
            if isinstance(ws2["{}{}".format(col, row)].value, str):
                l_name = ws2["{}{}".format(col, row)].value.lower()
            else:
                l_name = ws2["{}{}".format(col, row)].value = 'none'
            for person in peeps:
                if person.first_name == f_name and person.last_name == l_name:
                    #create a list of current members--to be used later#
                    new_person = Person(f_name, l_name)
                    members_set.add(new_person)
                    #populate cells with data#
                    ws2[initial] = person.initial
                    ws2[monthly] = person.monthly
                    ws2[hundreds] = person.hundreds
    return members_set

members_set = map_to_final()

def add_new_members():
    
    #compare sets for difference#        
    members_names = {member.first_name + ' ' + member.last_name for member in members_set}
    peeps_names = {person.first_name + ' ' + person.last_name for person in peeps}
    new_peeps = peeps_names.difference(members_names)    
    
    for person in peeps:
        for name in new_peeps:
            if person.first_name + ' ' + person.last_name == name:
                ws2["A" + str(ws2.max_row + 1)] = person.first_name
                ws2["B" + str(ws2.max_row)] = person.last_name
                ws2["G" + str(ws2.max_row)] = person.initial
                ws2["H" + str(ws2.max_row)] = person.monthly
                ws2["I" + str(ws2.max_row)] = person.hundreds
    
add_new_members()

def clean_up():
    print('Adding "zz" To Duplicates... Remember To Sort Member Sheet...')
    for row in range(1, ws2.max_row-1):
        for col in 'G':
            initial = ws2["{}{}".format(col, row)]
        for col in 'H':
            monthly = ws2["{}{}".format(col, row)]
        for col in 'I':
            hundreds = ws2["{}{}".format(col, row)]
        for col in 'A':
            f_cell = ws2["{}{}".format(col, row)]
            if isinstance(ws2["{}{}".format(col, row)].value, str):
                f_name = ws2["{}{}".format(col, row)].value.lower()
                next_f = ws2["{}{}".format(col, row+1)].value.lower()                
            else:
                l_name = ws2["{}{}".format(col, row)].value = 'none'
                next_l = ws2["{}{}".format(col, row+1)].value = 'none'
        for col in 'B':            
            if isinstance(ws2["{}{}".format(col, row)].value, str):
                l_name = ws2["{}{}".format(col, row)].value.lower()
                next_l = ws2["{}{}".format(col, row+1)].value.lower()
            else:
                l_name = ws2["{}{}".format(col, row)].value = 'none'
                next_l = ws2["{}{}".format(col, row+1)].value = 'none'           
            if f_name == next_f and l_name == next_l:
                f_cell.value = 'zz' + f_name
                initial.value = 0
                monthly.value = 0
                hundreds.value = 0

clean_up()
#input()
wb2.save('finished2_july_gtf.xlsx')
print('FINISHED')